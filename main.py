import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from pipeline.paginas import recorrer_paginas
from pipeline.detalles import recorrer_detalles


# ============================================================
# CONFIGURACIÓN DEL SCRAPER
# ============================================================
# La API empieza en 0:
# página 0 = primera página visible
# página 1 = segunda página visible
# página 2 = tercera página visible
#
# Ejemplos:
# desde = 0, hasta = 0  -> trae 1 sola página
# desde = 0, hasta = 1  -> trae 2 páginas
# desde = 0, hasta = 4  -> trae 5 páginas
DESDE_PAGINA = 0
HASTA_PAGINA = 3

# Cantidad máxima de cursos a procesar y guardar.
# Ejemplos:
# MAX_CURSOS = 3   -> guarda 3 cursos
# MAX_CURSOS = 10  -> guarda 10 cursos
# MAX_CURSOS = None -> guarda todos los cursos traídos
MAX_CURSOS = None

# Nombres de archivos de salida
NOMBRE_CSV = "cursos_wecandoo.csv"
NOMBRE_EXCEL = "cursos_wecandoo.xlsx"


def limpiar_valor_para_excel(valor):
    """
    Excel no acepta listas o diccionarios directamente.
    Esta función los convierte a texto.
    """

    if valor is None:
        return ""

    if isinstance(valor, (list, dict)):
        return json.dumps(valor, ensure_ascii=False)

    return valor


def guardar_csv(nombre_archivo, filas):
    """
    Guarda una lista de diccionarios en un archivo CSV.
    """

    if not filas:
        print("No hay datos para guardar en el CSV.")
        return

    columnas = list(filas[0].keys())

    with open(nombre_archivo, mode="w", newline="", encoding="utf-8-sig") as archivo:
        writer = csv.DictWriter(archivo, fieldnames=columnas)
        writer.writeheader()
        writer.writerows(filas)

    print(f"CSV guardado correctamente: {nombre_archivo}")


def guardar_excel(nombre_archivo, filas):
    """
    Guarda una lista de diccionarios en un archivo Excel (.xlsx)
    con formato simple y ordenado.
    """

    if not filas:
        print("No hay datos para guardar en el Excel.")
        return

    columnas = list(filas[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Cursos"

    # ------------------------------------------------------------
    # 1) Encabezados
    # ------------------------------------------------------------
    for col_num, nombre_columna in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_num, value=nombre_columna)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
        celda.alignment = Alignment(horizontal="center", vertical="center")

    # ------------------------------------------------------------
    # 2) Datos
    # ------------------------------------------------------------
    for fila_num, fila in enumerate(filas, start=2):
        for col_num, nombre_columna in enumerate(columnas, start=1):
            valor_original = fila.get(nombre_columna)
            valor_limpio = limpiar_valor_para_excel(valor_original)

            celda = ws.cell(row=fila_num, column=col_num, value=valor_limpio)
            celda.alignment = Alignment(vertical="center", wrap_text=False)

    # ------------------------------------------------------------
    # 3) Congelar primera fila
    # ------------------------------------------------------------
    ws.freeze_panes = "A2"

    # ------------------------------------------------------------
    # 4) Ajustar ancho de columnas
    # ------------------------------------------------------------
    for col_num, nombre_columna in enumerate(columnas, start=1):
        letra_columna = get_column_letter(col_num)
        largo_maximo = len(str(nombre_columna))

        for fila in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for celda in fila:
                if celda.value is not None:
                    largo_texto = len(str(celda.value))
                    if largo_texto > largo_maximo:
                        largo_maximo = largo_texto

        ancho_ajustado = largo_maximo + 2

        # Tope para que no queden columnas gigantes
        if ancho_ajustado > 50:
            ancho_ajustado = 50

        ws.column_dimensions[letra_columna].width = ancho_ajustado

    ws.row_dimensions[1].height = 24

    wb.save(nombre_archivo)

    print(f"Excel guardado correctamente: {nombre_archivo}")


def main():
    """
    Función principal del proyecto.
    """

    print("=== CONFIGURACIÓN ELEGIDA ===")
    print(f"Desde página: {DESDE_PAGINA}")
    print(f"Hasta página: {HASTA_PAGINA}")
    print(f"Máximo de cursos a guardar: {MAX_CURSOS}")
    print()

    # ------------------------------------------------------------
    # 1) Traemos cursos base desde el listado
    # ------------------------------------------------------------
    print("=== ETAPA 1: LEYENDO LISTADOS ===")
    cursos_base = recorrer_paginas(DESDE_PAGINA, HASTA_PAGINA)

    print(f"\nSe juntaron {len(cursos_base)} cursos base.")

    # ------------------------------------------------------------
    # 2) Recortamos la cantidad si MAX_CURSOS no es None
    # ------------------------------------------------------------
    if MAX_CURSOS is not None:
        cursos_base = cursos_base[:MAX_CURSOS]

    print(f"Se van a procesar {len(cursos_base)} cursos con detalle.\n")

    # ------------------------------------------------------------
    # 3) Recorremos el detalle de cada curso
    # ------------------------------------------------------------
    print("=== ETAPA 2: LEYENDO DETALLES ===")
    cursos_completos = recorrer_detalles(cursos_base)

    # ------------------------------------------------------------
    # 4) Guardamos archivos
    # ------------------------------------------------------------
    print("\n=== ETAPA 3: GUARDANDO ARCHIVOS ===")
    guardar_csv(NOMBRE_CSV, cursos_completos)
    guardar_excel(NOMBRE_EXCEL, cursos_completos)

    # ------------------------------------------------------------
    # 5) Resumen final
    # ------------------------------------------------------------
    print("\n=== RESUMEN FINAL ===")
    print(f"Total de cursos guardados: {len(cursos_completos)}")

    if cursos_completos:
        primer_curso = cursos_completos[0]

        print("\nPrimer curso guardado:")
        print("Nombre listado:", primer_curso.get("nom"))
        print("Precio:", primer_curso.get("prix"))
        print("Ciudad:", primer_curso.get("city_and_district"))
        print("URL corta:", primer_curso.get("page_url"))
        print("URL completa:", primer_curso.get("url_completa"))
        print("Título detalle:", primer_curso.get("titulo"))
        print("Descripción detalle:", primer_curso.get("descripcion"))


if __name__ == "__main__":
    main()